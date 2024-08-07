from django.shortcuts import render, redirect
from django.http import HttpResponse
from django.http import JsonResponse
from django.shortcuts import render
from openpyxl import load_workbook
from .models import *
from django.contrib.auth.decorators import user_passes_test
from django.core.paginator import Paginator
from django.core.mail import send_mail
from django.contrib import messages
from django.utils import timezone
from datetime import datetime

def is_ajax(request):
    return request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest'

def HomePageView (request):
    return render(request, 'index.html')

def MenuView(request):
    """
    Renders the menu view with categories and paginated menu items.
    
    This function retrieves all categories with their associated menu items and paginates the menu items with 9 items per page. 
    It checks if the request is AJAX and returns a JsonResponse with the paginated items and next page flag if so.
    Finally, it creates a context with categories and paginated menu items and renders the 'menu.html' template with this context.
    """
    categories = Category.objects.prefetch_related('menu_items').all()
    all_menu_items = MenuItem.objects.all()
    paginator = Paginator(all_menu_items, 9)  # Load 9 items at a time
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'categories': categories,
        'page_obj': page_obj,
    }
    return render(request, 'menu.html', context)


def AboutView(request):
    return render(request, 'about.html')

def BookView(request):
    """
    Handles table booking requests for the restaurant.

    This view processes the POST request from the booking form, validates the data,
    checks table availability, creates a reservation, and sends a notification email
    to the admin. It also handles errors and provides appropriate feedback to the user.

    Args:
        request (HttpRequest): The HTTP request object.

    Returns:
        HttpResponse: The HTTP response object with the booking page or a success message.
    """
    if request.method == 'POST':
        # Extract form data from the POST request
        customer_name = request.POST.get('customer_name')
        contact_info = request.POST.get('contact_info')
        email = request.POST.get('email')
        date_time = request.POST.get('date_time')
        number_of_people = request.POST.get('number_of_people')
        table_id = request.POST.get('table')

        try:
            # Parse the date and time from the form input and make it timezone-aware
            start_time = timezone.make_aware(datetime.strptime(date_time, "%Y-%m-%dT%H:%M"), timezone.get_current_timezone())
            duration = timedelta(hours=2)

            # Define restaurant operating hours (make them timezone-aware)
            restaurant_opening_time = timezone.make_aware(datetime.combine(start_time.date(), datetime.strptime("10:00", "%H:%M").time()), timezone.get_current_timezone())
            restaurant_closing_time = timezone.make_aware(datetime.combine(start_time.date(), datetime.strptime("22:00", "%H:%M").time()), timezone.get_current_timezone())

            # Validate the reservation time is within operating hours
            if not (restaurant_opening_time <= start_time <= restaurant_closing_time and 
                    restaurant_opening_time <= (start_time + duration) <= restaurant_closing_time):
                raise ValueError("Reservation time is outside of operating hours.")

            # Retrieve the selected table from the database
            table = Table.objects.get(id=table_id)

            # Check if the table is available for the requested time
            if table.is_reserved and table.reserved_until > start_time:
                raise ValueError("The selected table is not available for the requested time.")

            # Create a new reservation
            booking = Reservation.objects.create(
                customer_name=customer_name,
                contact_info=contact_info,
                email=email,
                date_time=start_time,
                table=table,
                status='Pending'
            )
            booking.save()

            # Update the table's reservation status
            table.is_reserved = True
            table.reserved_until = start_time + duration
            table.save()

            # Send an email notification to the admin
            subject = f'New Booking: {customer_name}'
            message_content = (
                f'Name: {customer_name}\n'
                f'Contact Info: {contact_info}\n'
                f'Email: {email}\n'
                f'Date and Time: {date_time}\n'
                f'Number of People: {number_of_people}\n'
                f'Table Group: {table.table_group}'
            )
            send_mail(
                subject,
                message_content,
                'your-email@example.com',
                ['admin@example.com'],
                fail_silently=False,
            )

            # Display success message and redirect to home page
            messages.success(request, 'Reservation successful! An admin will reach out to you to confirm the reservation.')
            return redirect('home')

        except Exception as e:
            # Handle any errors that occur and display an error message
            messages.error(request, f'Error saving booking: {str(e)}')
            return render(request, 'book.html', {'tables': Table.objects.all()})

    # Render the booking page with the available tables
    return render(request, 'book.html', {'tables': Table.objects.all()})


def event_list(request):
    """
    Display a list of events with their next occurrences.
    
    This view retrieves all events from the database, calculates the next occurrence
    for recurring events, and passes the events to the template for rendering.
    
    Args:
        request (HttpRequest): The HTTP request object.
        
    Returns:
        HttpResponse: The HTTP response object with the event list page.
    """
    events = Event.objects.prefetch_related('days').all()
    now = timezone.now()  

    for event in events:
        event.next_occurrence = event.get_next_occurrence()

    context = {
        'events': events
    }

    return render(request, 'events.html', context)




@user_passes_test(lambda u: u.is_superuser)
def import_from_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES['excel_file']
        wb = load_workbook(excel_file)
        ws = wb.active

        # Iterate starting from the second row (min_row=2)
        for row in ws.iter_rows(min_row=2, values_only=True): 
            try:
                category_name, item, description, price = row
                category, created = Category.objects.get_or_create(name=category_name)
                description = description or ""
                price = price or 0.00

                # create new menu item
                MenuItem.objects.create(
                    name=item,
                    description=description,
                    price=price,
                    category=category  
                )
            except ValueError:  
                print(f"Error processing row: {row}")

        return render(request, 'import_success.html')

    return render(request, 'import_form.html') 




# def import_from_excel(request, model, field_mapping, category_model=None, category_field=None):
#     """
#     Imports data from an Excel file and creates model instances.

#     Args:
#         request: The HTTP request object.
#         model: The Django model class to create instances for.
#         field_mapping: A dictionary mapping model fields to column indices in the Excel file.
#         category_model: (Optional) The Django model class for categories.
#         category_field: (Optional) The field in the data that represents the category.

#     Returns:
#         A rendered template response.
#     """
#     if request.method == 'POST':
#         # Retrieve the uploaded Excel file from the request
#         excel_file = request.FILES['excel_file']
#         # Load the workbook and get the active worksheet
#         wb = load_workbook(excel_file)
#         ws = wb.active

#         # Iterate over the rows in the worksheet, starting from the second row
#         for row in ws.iter_rows(min_row=2, values_only=True):
#             try:
#                 # Map row values to the model fields using the field_mapping
#                 data = {field: row[idx] for field, idx in field_mapping.items()}

#                 # If category information is provided, handle category creation/association
#                 if category_model and category_field:
#                     # Extract the category name from the data dictionary
#                     category_name = data.pop(category_field)
#                     # Get or create the category instance
#                     category, created = category_model.objects.get_or_create(name=category_name)
#                     # Add the category to the data dictionary
#                     data['category'] = category

#                 # Create a new instance of the model with the mapped data
#                 model.objects.create(**data)
#             except Exception as e:
#                 # Print an error message if there is an exception while processing the row
#                 print(f"Error processing row {row}: {e}")

#         # Render the success template after processing the Excel file
#         return render(request, 'import_success.html')

#     # Render the form template if the request method is not POST
#     return render(request, 'import_form.html')

# # Usage example
# def import_menu_items(request):
#     """
#     Handles the import of menu items from an Excel file.

#     Args:
#         request: The HTTP request object.

#     Returns:
#         A rendered template response from the import_from_excel function.
#     """
#     # Define the mapping of model fields to Excel column indices
#     field_mapping = {
#         'name': 1,          # Column B
#         'description': 2,   # Column C
#         'price': 3,         # Column D
#     }
#     # Call the import function with the MenuItem model and Category model
#     return import_from_excel(request, MenuItem, field_mapping, Category, 'category')